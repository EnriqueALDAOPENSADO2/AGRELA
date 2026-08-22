#!/usr/bin/env python3
import os
import sys
import json
import re
import glob
import xlrd
import openpyxl

OUTPUT_DIR = "extracted_images"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def clean_code(val):
    if val is None or val == "":
        return ""
    if isinstance(val, float):
        if val.is_integer():
            return str(int(val))
    s = str(val).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        return s[:-2]
    return s

def clean_float(val):
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip().replace("€", "").replace(" ", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def build_smart_image_lookup():
    all_imgs = sorted(glob.glob(os.path.join(OUTPUT_DIR, "*.png")))
    
    def get_image(code, desc, category=""):
        if code:
            for img in all_imgs:
                base = os.path.basename(img)
                if code in base:
                    return img
        if code and len(code) >= 6:
            prefix = code[:6]
            for img in all_imgs:
                base = os.path.basename(img)
                if prefix in base:
                    return img
        d = desc.lower()
        if "cajon blanco" in d:
            return "extracted_images/compactos_290155000.png"
        if "lama 43" in d or "aluminio de 43" in d:
            return "extracted_images/lam2026_t0_280043008.png"
        if "pvc r - 45" in d or "pvc radial 45" in d:
            return "extracted_images/lam2026_t1_220010001.png"
        if "pvc mini" in d:
            return "extracted_images/lam2026_t2_260010001.png"
        if "bajera de 44" in d:
            return "extracted_images/lam2026_t3_280000015.png"
        if "bajera de 53" in d:
            return "extracted_images/lam2026_t4_280000019.png"
        if "juego de perfiles 155" in d:
            return "extracted_images/lam2026_t5_290000001.png"
        if "juego de perfiles 185" in d:
            return "extracted_images/lam2026_t5_290000004.png"
        if "perfil lateral" in d:
            return "extracted_images/lam2026_t6_290000007.png"
        if "perfil fondo" in d:
            return "extracted_images/lam2026_t7_290000011.png"
        if "autoblocante" in d:
            return "extracted_images/compactos_290155007.png"
        if "motor" in d:
            return "extracted_images/motor_p1_600500001.png"
        if "mosquitera" in d or "mosquiflex" in d:
            return "extracted_images/motor_p1_600500037.png"
        return ""

    return get_image

def get_mosquiflex_catalog():
    items_raw = [
        # Mosquiteras Enrollables Ventana 42mm
        ("MQ-ENR-01", "Mosquitera Enrollable Ventana 42 - Grupo 1 (Blanco)", 48.00, "ud.", "Enrollables Ventana"),
        ("MQ-ENR-02", "Mosquitera Enrollable Ventana 42 - Grupo 2 (Plata / Bronce / RAL Estándar)", 54.00, "ud.", "Enrollables Ventana"),
        ("MQ-ENR-03", "Mosquitera Enrollable Ventana 42 - Grupo 3 (Madera / Foliado)", 62.00, "ud.", "Enrollables Ventana"),
        ("MQ-ENR-MUE", "Muelle de retención para Mosquitera Enrollable 42", 12.00, "ud.", "Accesorios y Componentes"),
        ("MQ-ENR-AV1", "Guías con 1 Felpudo Antiviento (Incremento)", 3.80, "ud.", "Accesorios y Componentes"),
        ("MQ-ENR-AV2", "Guías con 2 Felpudos Antiviento (Incremento)", 7.60, "ud.", "Accesorios y Componentes"),
        ("MQ-ENR-DAV1", "Guía doble con 1 Felpudo Antiviento (Incremento)", 3.80, "ud.", "Accesorios y Componentes"),
        ("MQ-ENR-DAV2", "Guía doble con 2 Felpudos Antiviento (Incremento)", 7.60, "ud.", "Accesorios y Componentes"),
        ("MQ-ENR-TET20", "Malla TETUG 20-1400 enrollada en eje", 28.50, "ud.", "Mallas y Tejidos"),
        ("MQ-ENR-TET25", "Malla TETUG 25-1400 enrollada en eje", 33.20, "ud.", "Mallas y Tejidos"),
        ("MQ-ENR-TET23", "Malla TETUG 23-1400 Enrollable Puerta", 37.12, "ud.", "Mallas y Tejidos"),

        # Mosquitera Enrollable Puerta Lateral
        ("MQ-ENR-PTA1", "Mosquitera Enrollable Puerta Lateral Única - Grupo 1", 125.00, "ud.", "Enrollable Puerta"),
        ("MQ-ENR-PTA2", "Mosquitera Enrollable Puerta Lateral Doble - Grupo 1", 230.00, "ud.", "Enrollable Puerta"),
        ("MQ-CMP-PTA01", "Perfil Cierre Imán Enrollable Puerta", 11.47, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA02", "Eslabón Antiviento Enrollable Puerta (Juego)", 7.56, "ud.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA03", "Cadena Inferior Enrollable Puerta", 49.08, "ud.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA04", "Imán Perfil Móvil Enrollable Puerta", 2.20, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA05", "Imán Perfil Cierre Enrollable Puerta", 2.76, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA06", "Portaimán Enrollable Puerta", 5.77, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA07", "Hilo Trenzado Negro Enrollable Puerta", 1.88, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA08", "Carril Inferior Enrollable Puerta", 14.57, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA09", "Perfil Móvil Enrollable Puerta (G1 Blanco)", 38.18, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA10", "Carril Superior Enrollable Puerta (G1 Blanco)", 28.84, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA11", "Perfil Fijo Enrollable Puerta (G1 Blanco)", 31.97, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA12", "Fleje Enrollable Puerta", 7.00, "ml.", "Componentes Enrollable Puerta"),
        ("MQ-CMP-PTA13", "Kit Felpudo 4,8 x 19 mm Enrollable Puerta", 85.88, "ud.", "Componentes Enrollable Puerta"),

        # Mosquitera Plisada 22
        ("MQ-PLI-22-1", "Mosquitera Plisada 22 Puerta Única - Grupo 1", 135.00, "ud.", "Plisada 22"),
        ("MQ-PLI-22-2", "Mosquitera Plisada 22 Puerta Doble - Grupo 1", 245.00, "ud.", "Plisada 22"),
        ("MQ-PLI-22-REV", "Mosquitera Plisada 22 Reversible - Grupo 1", 155.00, "ud.", "Plisada 22"),
        ("MQ-PLI-22-VEN", "Mosquitera Plisada 22 Ventana - Grupo 1", 85.00, "ud.", "Plisada 22"),
        ("MQ-CMP-PL22-01", "Perfil Fijo Plisada 22 (G1 Blanco)", 12.37, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-02", "Carril Superior Plisada 22 (G1 Blanco)", 13.64, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-03", "Perfil Móvil Plisada 22 (G1 Blanco)", 21.98, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-04", "Carril Inferior Aluminio Negro Plisada 22", 6.95, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-05", "Adhesivo Carril Inferior Plisada 22", 2.00, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-06", "Clip Hilo Plisada 22 / 40", 1.04, "ud.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-07", "Hilo Negro Plisada 22 / 40", 1.88, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-08", "Imán Perfil Fijo Plisada 22", 2.97, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-09", "Imán Perfil Móvil Plisada 22", 2.20, "ml.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-10", "Malla Plisada 22 - 2400 mm", 50.64, "ud.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-11", "Malla Plisada 22 - 3000 mm", 77.40, "ud.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-12", "Lámina Fijación Malla 22", 3.88, "ud.", "Componentes Plisada 22"),
        ("MQ-CMP-PL22-13", "Perfil Instalación Frontal Plisada 22 (G1 Blanco)", 22.95, "ml.", "Componentes Plisada 22"),

        # Mosquitera Plisada 40
        ("MQ-PLI-40-1", "Mosquitera Plisada 40 Puerta Única - Grupo 1", 326.00, "ud.", "Plisada 40"),
        ("MQ-PLI-40-2", "Mosquitera Plisada 40 Puerta Doble - Grupo 1", 653.00, "ud.", "Plisada 40"),
        ("MQ-CMP-PL40-01", "Perfil Fijo Plisada 40 (G1 Blanco)", 23.90, "ml.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-02", "Perfil Móvil Plisada 40 (G1 Blanco)", 24.90, "ml.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-03", "Carril Superior Plisada 40 (G1 Blanco)", 19.10, "ml.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-04", "Perfil Cierre Imán Plisada 40", 11.93, "ml.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-05", "Porta Imán Plisada 40", 4.43, "ml.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-06", "Carril Inferior Gris Adhesivo Plisada 40", 29.28, "ml.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-07", "Cadena 58 Eslabones Plisada 40", 48.60, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-08", "Cadena 3 Eslabones Plisada 40", 5.80, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-09", "Eslabón Reversible Plisada 40", 1.40, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-10", "Eslabón Gancho Plisada 40", 1.65, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-11", "Eslabón Rojo Plisada 40", 2.12, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-12", "Eslabón Blanco Plisada 40", 2.00, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-13", "Maneta Plisada 40", 2.40, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-14", "Malla Plisada 40 - 2200 mm", 56.04, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-15", "Malla Plisada 40 - 2600 mm", 62.56, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-16", "Cantonera Superior P. Móvil 40", 8.92, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-17", "Cantonera Superior P. Fijo 40", 7.92, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-18", "Cantonera Inferior P. Móvil 40", 8.44, "ud.", "Componentes Plisada 40"),
        ("MQ-CMP-PL40-19", "Cantonera Inferior P. Fijo 40", 7.92, "ud.", "Componentes Plisada 40"),

        # Mosquiteras Abatibles, Fijas, Correderas y Formas
        ("MQ-ABA-1", "Mosquitera Abatible Puerta 1 Hoja - Grupo 1 (Blanco)", 303.00, "ud.", "Abatible Puerta"),
        ("MQ-ABA-2", "Mosquitera Abatible Puerta 2 Hojas - Grupo 1 (Blanco)", 605.00, "ud.", "Abatible Puerta"),
        ("MQ-CMP-AB01", "Perfil Marco Abatible (G1 Blanco)", 18.50, "ml.", "Componentes Abatible"),
        ("MQ-CMP-AB02", "Perfil Hoja Abatible (G1 Blanco)", 21.20, "ml.", "Componentes Abatible"),
        ("MQ-CMP-AB03", "Perfil Encuentro Abatible 2 Hojas (G1 Blanco)", 13.27, "ml.", "Componentes Abatible"),
        ("MQ-CMP-AB04", "Perfil Travesaño Abatible (G1 Blanco)", 10.93, "ml.", "Componentes Abatible"),
        ("MQ-CMP-AB05", "Bisagra con Muelle de Retorno Abatible", 12.80, "ud.", "Componentes Abatible"),
        ("MQ-FIJ-1", "Mosquitera Fija con Marco - Grupo 1 (Blanco)", 38.00, "m²", "Mosquitera Fija"),
        ("MQ-FIJ-2", "Mosquitera Fija con Marco - Grupo 2 (Plata / RAL)", 44.00, "m²", "Mosquitera Fija"),
        ("MQ-COR-1", "Mosquitera Corredera Perfil Curvo - Grupo 1 (Blanco)", 45.00, "m²", "Mosquitera Corredera"),
        ("MQ-COR-2", "Mosquitera Corredera Perfil Curvo - Grupo 2 (Plata / RAL)", 52.00, "m²", "Mosquitera Corredera"),
        ("MQ-CMP-CR01", "Marco Corredera Guía Z - Grupo 1 (Blanco)", 18.60, "ml.", "Componentes Corredera"),
        ("MQ-CMP-CR02", "Rodamientos Rueda Corredera (Juego)", 4.80, "ud.", "Componentes Corredera"),
        ("MQ-INC-TRAV", "Travesaño Adicional Mosquitera (Incremento)", 23.80, "ud.", "Incrementos y Formas"),
        ("MQ-INC-IRR", "Forma Irregular / Plantilla (Incremento)", 27.00, "ud.", "Incrementos y Formas"),
        ("MQ-INC-MEDP", "Medio Punto / Arco Curva Monoradio (Incremento)", 100.00, "ud.", "Incrementos y Formas"),
        ("MQ-INC-BUEY", "Ojo de Buey / Curva Multiradio (Incremento)", 250.00, "ud.", "Incrementos y Formas"),
        ("MQ-EXP-01", "Carta de Colores Mosquiflex", 10.00, "ud.", "Muestras y Expositores"),
        ("MQ-EXP-02", "Expositor Personalizable Mosquiflex", 590.00, "ud.", "Muestras y Expositores")
    ]

    items = []
    for code, desc, pvp, unit, cat in items_raw:
        items.append({
            "sheet": "Mosquiflex",
            "category": cat,
            "code": code,
            "desc": desc,
            "pvp": pvp,
            "t1": round(pvp * 0.75, 2),
            "u1": unit,
            "p1": pvp,
            "p_t1": round(pvp * 0.75, 2),
            "img_path": "extracted_images/motor_p1_600500037.png"
        })
    return items

def get_flexol_catalog():
    items_raw = [
        # Venecianas de Aluminio y Madera
        ("FLX-VEN-16", "Veneciana Aluminio 16 mm - Colores Básicos", 35.00, "m²", "Venecianas"),
        ("FLX-VEN-25", "Veneciana Aluminio 25 mm - Colores Básicos", 29.00, "m²", "Venecianas"),
        ("FLX-VEN-50", "Veneciana Aluminio 50 mm - Colores Básicos con Cordón", 42.00, "m²", "Venecianas"),
        ("FLX-VEN-50C", "Veneciana Aluminio 50 mm - Colores Básicos con Cinta", 48.00, "m²", "Venecianas"),
        ("FLX-VEN-MAD", "Veneciana Madera 50 mm - Colección Madeira / Bali", 78.00, "m²", "Venecianas Madera"),
        ("FLX-VEN-MAD2", "Veneciana Madera 50 mm - Colección Ceylan / Seychelles", 88.00, "m²", "Venecianas Madera"),
        ("FLX-CMP-VN01", "Cabezal de Aluminio 25 mm", 4.86, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN02", "Cabezal de Aluminio 50 mm", 7.36, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN03", "Terminal Inferior Aluminio 25 mm", 2.04, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN04", "Terminal Inferior Aluminio 50 mm", 5.52, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN05", "Galería Madera G-1 para incrustar en hueco", 15.51, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN06", "Galería Madera G-3 en U", 29.42, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN07", "Maquinilla Freno Cordón 25 mm", 2.17, "ud.", "Componentes Venecianas"),
        ("FLX-CMP-VN08", "Maquinilla Orientador Varilla 25 mm", 1.78, "ud.", "Componentes Venecianas"),
        ("FLX-CMP-VN09", "Escalerilla Cordón 25 mm", 0.40, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN10", "Escalerilla Cinta 50 mm", 1.80, "ml.", "Componentes Venecianas"),
        ("FLX-CMP-VN11", "Soporte Techo / Pared 25 mm (Juego)", 1.20, "ud.", "Componentes Venecianas"),
        ("FLX-CMP-VN12", "Soporte Techo / Pared 50 mm (Juego)", 2.40, "ud.", "Componentes Venecianas"),
        ("FLX-CMP-VN13", "Motor Somfy para Veneciana 24V / 230V", 145.00, "ud.", "Componentes Venecianas"),

        # Cortinas Verticales (89mm y 127mm)
        ("FLX-VT-89", "Cortina Vertical 89 mm - Colección Teide / Alcazaba", 42.00, "m²", "Cortinas Verticales"),
        ("FLX-VT-127", "Cortina Vertical 127 mm - Colección Teide / Alcazaba", 38.00, "m²", "Cortinas Verticales"),
        ("FLX-VT-89M", "Cortina Vertical 89 mm - Colección Mulhacén / Aneto", 49.00, "m²", "Cortinas Verticales"),
        ("FLX-VT-127M", "Cortina Vertical 127 mm - Colección Mulhacén / Aneto", 44.00, "m²", "Cortinas Verticales"),
        ("FLX-VT-MAD", "Cortina Vertical Madera 89 mm", 111.80, "m²", "Cortinas Verticales"),
        ("FLX-VT-CUE", "Cortina Vertical Cuero 89 mm", 141.85, "m²", "Cortinas Verticales"),
        ("FLX-CMP-VT01", "Riel Vertical Aluminio Blanco / Plata", 17.40, "ml.", "Componentes Verticales"),
        ("FLX-CMP-VT02", "Carro Arrastre Vertical con Gancho", 0.59, "ud.", "Componentes Verticales"),
        ("FLX-CMP-VT03", "Cadena de Distancia Inferior 89 mm", 0.35, "ml.", "Componentes Verticales"),
        ("FLX-CMP-VT04", "Cadena de Distancia Inferior 127 mm", 0.45, "ml.", "Componentes Verticales"),
        ("FLX-CMP-VT05", "Contrapeso Inferior 89 mm", 0.76, "ud.", "Componentes Verticales"),
        ("FLX-CMP-VT06", "Contrapeso Inferior 127 mm", 0.95, "ud.", "Componentes Verticales"),
        ("FLX-CMP-VT07", "Motor para Cortina Vertical 230V con Mando", 365.00, "ud.", "Componentes Verticales"),

        # Cortinas Plisadas
        ("FLX-PLI-01", "Cortina Plisada Premium PLI.01 - Colección 236", 58.00, "m²", "Cortinas Plisadas"),
        ("FLX-PLI-REV", "Cortina Plisada Repliegue Reversible PLI.02", 68.00, "m²", "Cortinas Plisadas"),
        ("FLX-PLI-ND", "Cortina Plisada Noche y Día PLI.03", 85.00, "m²", "Cortinas Plisadas"),
        ("FLX-PLI-PLI", "Cortina Plisada Plano Inclinado / Buhardilla PLI.04", 74.00, "m²", "Cortinas Plisadas"),
        ("FLX-PLI-MONO", "Cortina Plisada Plus Monocomando PLI.05", 79.00, "m²", "Cortinas Plisadas"),
        ("FLX-PLI-MINI", "Cortina Plisada Mini Cristal Accionamiento Manual", 49.00, "m²", "Cortinas Plisadas"),
        ("FLX-PLI-MINIC", "Cortina Plisada Mini Accionamiento Cordón", 45.00, "m²", "Cortinas Plisadas"),
        ("FLX-CMP-PL01", "Perfil Plisada Mini Cristal Blanco (13,3 x 20 mm)", 19.15, "ml.", "Componentes Plisadas"),
        ("FLX-CMP-PL02", "Perfil Fijo Mini Cristal", 15.81, "ml.", "Componentes Plisadas"),
        ("FLX-CMP-PL03", "Soporte Tensor Mini Modelo B", 0.50, "ud.", "Componentes Plisadas"),
        ("FLX-CMP-PL04", "Freno Cordón Plisada Mini", 1.70, "ud.", "Componentes Plisadas"),
        ("FLX-CMP-PL05", "Cordoncillo Plisada 1,2 mm", 0.63, "ml.", "Componentes Plisadas"),

        # Cortinas Plegables y Paneles Deslizantes
        ("FLX-PLE-01", "Cortina Plegable Confeccionada Sistema Plus PLE.01", 55.00, "m²", "Cortinas Plegables"),
        ("FLX-PLE-LON", "Cortina Plegable La Loneta PLE.05", 62.00, "m²", "Cortinas Plegables"),
        ("FLX-PLE-VIS", "Cortina Plegable El Visillo PLE.06", 58.00, "m²", "Cortinas Plegables"),
        ("FLX-CMP-PLE01", "Mecanismo Plegable a Cadena con Reductor", 22.50, "ud.", "Componentes Plegables"),
        ("FLX-CMP-PLE02", "Perfil Velcrado Plegable Plus", 10.92, "ml.", "Componentes Plegables"),
        ("FLX-CMP-PLE03", "Pletina Inferior Aluminio Plegable 40 mm", 3.55, "ml.", "Componentes Plegables"),
        ("FLX-CMP-PLE04", "Varilla Fibra de Vidrio Plegable 4 mm", 0.60, "ml.", "Componentes Plegables"),
        ("FLX-PAN-2V", "Panel Deslizante / Japonés 2 Vías Sistema Básico", 48.00, "ml.", "Paneles Deslizantes"),
        ("FLX-PAN-3V", "Panel Deslizante / Japonés 3 Vías Sistema Básico", 65.00, "ml.", "Paneles Deslizantes"),
        ("FLX-PAN-4V", "Panel Deslizante / Japonés 4 Vías Sistema Básico", 82.00, "ml.", "Paneles Deslizantes"),
        ("FLX-PAN-5V", "Panel Deslizante / Japonés 5 Vías Sistema Básico", 98.00, "ml.", "Paneles Deslizantes"),
        ("FLX-PAN-MOT", "Panel Deslizante Sistema Motor 3 a 5 Vías", 280.00, "ud.", "Paneles Deslizantes"),
        ("FLX-CMP-PAN01", "Portatela Velcrado Deslizante", 19.77, "ml.", "Componentes Paneles"),
        ("FLX-CMP-PAN02", "Trineo Arrastre Panel Deslizante", 7.51, "ud.", "Componentes Paneles"),
        ("FLX-CMP-PAN03", "Varilla de Mando Tirador 1,50 m", 3.38, "ud.", "Componentes Paneles"),

        # Cortinas Enrollables Flexol
        ("FLX-ENR-UNI", "Cortina Enrollable Universal EN.0 (Sin Tejido)", 28.00, "ud.", "Enrollables Flexol"),
        ("FLX-ENR-SAT", "Cortina Enrollable Screen Satiné 5500", 63.00, "m²", "Enrollables Flexol"),
        ("FLX-ENR-KAR", "Cortina Enrollable Screen Karellis 11301", 68.00, "m²", "Enrollables Flexol"),
        ("FLX-ENR-OPA", "Cortina Enrollable Tejido Opaco Vela", 54.00, "m²", "Enrollables Flexol"),
        ("FLX-ENR-BAL", "Cortina Enrollable Tejido Bali / Madeira EN.06", 72.00, "m²", "Enrollables Flexol"),
        ("FLX-ENR-CAJ42", "Cortina Enrollable con Cajón de 42 mm", 72.00, "m²", "Enrollables con Cajón"),
        ("FLX-ENR-CAJ90", "Cortina Enrollable con Cajón de 90 mm CA.04", 98.00, "m²", "Enrollables con Cajón"),
        ("FLX-ENR-ZIP", "Cortina Enrollable Sistema ZIP Guiada", 135.00, "m²", "Enrollables Guiadas ZIP"),
        ("FLX-ENR-CAB", "Cortina Enrollable Guiada a Cable de Acero", 78.00, "m²", "Enrollables Guiadas"),
        ("FLX-ENR-BUH", "Cortina Enrollable Buhardilla con Guías Laterales", 85.00, "m²", "Enrollables Especiales"),
        ("FLX-LM-01", "Cortina Luz Mágica LM.01 Alborada Galería 80", 88.00, "m²", "Cortinas Luz Mágica"),
        ("FLX-LM-03", "Cortina Luz Mágica LM.03 Niebla Galería 90", 96.00, "m²", "Cortinas Luz Mágica"),
        ("FLX-CMP-EN01", "Eje Aluminio D.43 mm para Enrollable", 9.76, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN02", "Eje Aluminio D.58 mm para Enrollable", 19.92, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN03", "Eje Aluminio D.80 mm para Enrollable", 34.52, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN04", "Mecanismo Cadena D.43 Desmultiplicado 1:3", 19.06, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN05", "Mecanismo Cadena D.58 Desmultiplicado 1:4", 39.26, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN06", "Cadena Mando PVC 4,5 x 6 mm Cerrada 1,50 m", 4.45, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN07", "Cadena Mando PVC 4,5 x 6 mm Cerrada 2,00 m", 5.34, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN08", "Cadena Mando Metálica Continua", 6.50, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN09", "Soportes Pared / Techo Enrollable 40 mm (Juego)", 5.00, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN10", "Soportes Pared / Techo Enrollable 60 mm (Juego)", 7.31, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN11", "Soportes Pared / Techo Enrollable 75 mm (Juego)", 14.49, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN12", "Cajón de Aluminio 90 mm Curvo / Recto", 33.51, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN13", "Guía ZIP 50 x 34 mm con Clip", 22.60, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN14", "Guía Ventilada Blanca con Pistón", 8.64, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN15", "Guía Cerrada 55 mm Instalación en Hueco / Frontal", 19.92, "ml.", "Componentes Enrollables"),
        ("FLX-CMP-EN16", "Accionamiento a Manivela Máquina 1:8 con Manivela 1,50 m", 205.00, "ud.", "Componentes Enrollables"),
        ("FLX-CMP-EN17", "Motorización Somfy para Enrollable D.40 / D.50", 165.00, "ud.", "Componentes Enrollables")
    ]

    items = []
    for code, desc, pvp, unit, cat in items_raw:
        items.append({
            "sheet": "Flexol",
            "category": cat,
            "code": code,
            "desc": desc,
            "pvp": pvp,
            "t1": round(pvp * 0.75, 2),
            "u1": unit,
            "p1": pvp,
            "p_t1": round(pvp * 0.75, 2),
            "img_path": "extracted_images/motor_p1_600500001.png"
        })
    return items

def extract_all_catalog_precios(precios_dir="PRECIOS", output_json="catalog.json"):
    print(f"Extrayendo catálogo completo 2026 + Flexol + Mosquiflex desde: {precios_dir}")
    
    get_img = build_smart_image_lookup()
    products_map = {}

    # 1. ACCESORIOS ENERO 2026.xls
    acc_files = [os.path.join(precios_dir, "ACCESORIOS ENERO 2026.xls"), os.path.join(precios_dir, "accesorios.xls")]
    acc_xls = next((f for f in acc_files if os.path.exists(f)), None)
    if acc_xls:
        wb = xlrd.open_workbook(acc_xls)
        for sname, sheet_name, default_cat in [
            ("ACCESORIO-ENRO PVP", "Accesorios", "Accesorios Persiana Enrollable"),
            ("ACCESORIO-COMP PVP", "Accesorios", "Accesorios Compactos")
        ]:
            if sname in wb.sheet_names():
                sh = wb.sheet_by_name(sname)
                cur_cat = default_cat
                for r in range(sh.nrows):
                    c0 = str(sh.cell_value(r, 0)).strip()
                    if "ACCESORIOS" in c0: cur_cat = c0
                    code = clean_code(sh.cell_value(r, 1))
                    desc = str(sh.cell_value(r, 2)).strip()
                    pvp = clean_float(sh.cell_value(r, 3))
                    unit = str(sh.cell_value(r, 4)).strip() or "ud."
                    if code and desc and pvp > 0:
                        products_map[code] = {
                            "sheet": sheet_name, "category": cur_cat, "code": code,
                            "desc": desc, "pvp": pvp, "t1": pvp, "u1": unit,
                            "p1": pvp, "p_t1": pvp, "img_path": get_img(code, desc, cur_cat)
                        }
        for sname in ["ACCESORIO-ENRO T-1", "ACCESORIO-COMP T-1"]:
            if sname in wb.sheet_names():
                sh = wb.sheet_by_name(sname)
                for r in range(sh.nrows):
                    code = clean_code(sh.cell_value(r, 1))
                    t1 = clean_float(sh.cell_value(r, 3))
                    if code in products_map and t1 > 0:
                        products_map[code]["t1"] = t1
                        products_map[code]["p_t1"] = t1

    # 2. ENROLLABLES ENERO 2026.xls
    enr_files = [os.path.join(precios_dir, "ENROLLABLES ENERO 2026.xls"), os.path.join(precios_dir, "enrollables.xls")]
    enr_xls = next((f for f in enr_files if os.path.exists(f)), None)
    if enr_xls:
        wb = xlrd.open_workbook(enr_xls)
        if "TARIFA PVP" in wb.sheet_names():
            sh = wb.sheet_by_name("TARIFA PVP")
            cur_cat = "Persianas Enrollables"
            for r in range(sh.nrows):
                c0 = str(sh.cell_value(r, 0)).strip()
                c1 = str(sh.cell_value(r, 1)).strip()
                if "PERSIANA" in c1 or "PERSIANA" in c0: cur_cat = c1 or c0
                code = clean_code(sh.cell_value(r, 1))
                desc = str(sh.cell_value(r, 2)).strip()
                pvp = clean_float(sh.cell_value(r, 3))
                unit = str(sh.cell_value(r, 4)).strip() or "m²"
                if code and code.isdigit() and desc and pvp > 0:
                    products_map[code] = {
                        "sheet": "Persianas Enrollables", "category": cur_cat, "code": code,
                        "desc": desc, "pvp": pvp, "t1": pvp, "u1": unit,
                        "p1": pvp, "p_t1": pvp, "img_path": get_img(code, desc, cur_cat)
                    }
        if "TARIFA T-1" in wb.sheet_names():
            sh = wb.sheet_by_name("TARIFA T-1")
            for r in range(sh.nrows):
                code = clean_code(sh.cell_value(r, 1))
                t1 = clean_float(sh.cell_value(r, 3))
                if code in products_map and t1 > 0:
                    products_map[code]["t1"] = t1
                    products_map[code]["p_t1"] = t1

    # 3. LAMAS Y CAJONES 2026.xls
    lam_files = [os.path.join(precios_dir, "LAMAS Y CAJONES 2026.xls"), os.path.join(precios_dir, "lamas_y_cajones.xls")]
    lam_xls = next((f for f in lam_files if os.path.exists(f)), None)
    if lam_xls:
        wb = xlrd.open_workbook(lam_xls)
        for pvp_sheet, t1_sheet, sheet_title, default_cat in [
            ("GUIAS y PERFILES PVP", "GUIAS y PERFILES T-1", "Guías y Perfiles", "Guías y Perfiles"),
            ("LAMAS y PERFILES PVP", "LAMAS y PERFILES T-1", "Lamas y Cajones", "Lamas y Perfiles")
        ]:
            if pvp_sheet in wb.sheet_names():
                sh = wb.sheet_by_name(pvp_sheet)
                cur_cat = default_cat
                for r in range(sh.nrows):
                    c0 = str(sh.cell_value(r, 0)).strip()
                    if "GUIAS" in c0 or "PERFILES" in c0 or "LAMAS" in c0: cur_cat = c0
                    code = clean_code(sh.cell_value(r, 1))
                    desc = str(sh.cell_value(r, 2)).strip()
                    pvp = clean_float(sh.cell_value(r, 3))
                    unit = str(sh.cell_value(r, 4)).strip() or "ml."
                    if (code and (code.isdigit() or "Tapón" in desc)) and desc and pvp > 0:
                        products_map[code] = {
                            "sheet": sheet_title, "category": cur_cat, "code": code,
                            "desc": desc, "pvp": pvp, "t1": pvp, "u1": unit,
                            "p1": pvp, "p_t1": pvp, "img_path": get_img(code, desc, cur_cat)
                        }
            if t1_sheet in wb.sheet_names():
                sh = wb.sheet_by_name(t1_sheet)
                for r in range(sh.nrows):
                    code = clean_code(sh.cell_value(r, 1))
                    t1 = clean_float(sh.cell_value(r, 3))
                    if code in products_map and t1 > 0:
                        products_map[code]["t1"] = t1
                        products_map[code]["p_t1"] = t1

    # 4. TARIFAS MOTORES ENERO 2026.xls
    mot_files = [os.path.join(precios_dir, "TARIFAS MOTORES ENERO 2026.xls"), os.path.join(precios_dir, "motores.xls")]
    mot_xls = next((f for f in mot_files if os.path.exists(f)), None)
    if mot_xls:
        wb = xlrd.open_workbook(mot_xls)
        if "MOTORES TARIFA PVP" in wb.sheet_names():
            sh = wb.sheet_by_name("MOTORES TARIFA PVP")
            cur_cat = "Motores y Automatismos"
            for r in range(sh.nrows):
                c0 = str(sh.cell_value(r, 0)).strip()
                if "MOTORES" in c0: cur_cat = c0
                code = clean_code(sh.cell_value(r, 1))
                desc = str(sh.cell_value(r, 2)).strip()
                pvp = clean_float(sh.cell_value(r, 3))
                unit = str(sh.cell_value(r, 4)).strip() or "ud."
                if code and code.isdigit() and desc and pvp > 0:
                    products_map[code] = {
                        "sheet": "Motores y Automatismos", "category": cur_cat, "code": code,
                        "desc": desc, "pvp": pvp, "t1": pvp, "u1": unit,
                        "p1": pvp, "p_t1": pvp, "img_path": get_img(code, desc, cur_cat)
                    }
        if "MOTORES TARIFA T-1" in wb.sheet_names():
            sh = wb.sheet_by_name("MOTORES TARIFA T-1")
            for r in range(sh.nrows):
                code = clean_code(sh.cell_value(r, 1))
                t1 = clean_float(sh.cell_value(r, 3))
                if code in products_map and t1 > 0:
                    products_map[code]["t1"] = t1
                    products_map[code]["p_t1"] = t1

    # 5. VENECIANAS ENERO 2026.xlsx
    ven_files = [os.path.join(precios_dir, "VENECIANAS ENERO 2026.xlsx"), os.path.join(precios_dir, "venecianas.xlsx")]
    ven_xlsx = next((f for f in ven_files if os.path.exists(f)), None)
    if ven_xlsx:
        wb = openpyxl.load_workbook(ven_xlsx, data_only=True)
        ws = wb.active
        v_idx = 1
        for r in range(1, ws.max_row + 1):
            val = ws.cell(r, 1).value
            if val:
                val_str = str(val).strip().replace("45 00 €", "45,00 €")
                m = re.search(r"^(.*?)\s+([0-9]+(?:[\.,][0-9]+)?)\s*€", val_str)
                if m:
                    desc = m.group(1).strip()
                    pvp = float(m.group(2).replace(",", "."))
                    t1 = round(pvp * 0.75, 2)
                    unit = "m²" if "GRADUABLE" in desc else ("ml." if "CORDON" in desc or "ESCALERILLA" in desc else "ud.")
                    code = f"VEN{v_idx:03d}"
                    v_idx += 1
                    products_map[code] = {
                        "sheet": "Venecianas y Graduables", "category": "Venecianas", "code": code,
                        "desc": desc, "pvp": pvp, "t1": t1, "u1": unit,
                        "p1": pvp, "p_t1": t1, "img_path": get_img(code, desc, "Venecianas")
                    }

    # 6. MOSQUIFLEX
    for it in get_mosquiflex_catalog():
        products_map[it["code"]] = it

    # 7. FLEXOL
    for it in get_flexol_catalog():
        products_map[it["code"]] = it

    all_products = list(products_map.values())
    with open(output_json, "w", encoding="utf-8") as out:
        json.dump(all_products, out, ensure_ascii=False, indent=2)

    # Generar precios_catalogo.xlsx
    wb_cat = openpyxl.Workbook()
    ws_cat = wb_cat.active
    ws_cat.title = "Precios"
    ws_cat.append(["Código", "Descripción", "PVP (€)", "Tarifa 1 (€)", "Unidad", "Categoría / Hoja", "Croquis Imagen"])

    for it in all_products:
        ws_cat.append([
            it.get("code", ""),
            it.get("desc", ""),
            it.get("pvp", it.get("p1", 0.0)),
            it.get("t1", it.get("p_t1", 0.0)),
            it.get("u1", "ud."),
            it.get("sheet", ""),
            it.get("img_path", "")
        ])

    wb_cat.save("precios_catalogo.xlsx")
    wb_cat.save("windows/precios_catalogo.xlsx")
    with open("windows/catalog.json", "w", encoding="utf-8") as out_win:
        json.dump(all_products, out_win, ensure_ascii=False, indent=2)

    print(f"Catálogo completo generado con {len(all_products)} artículos limpios en '{output_json}' y 'precios_catalogo.xlsx'.")
    return all_products

if __name__ == "__main__":
    p_dir = sys.argv[1] if len(sys.argv) > 1 else "PRECIOS"
    out_f = sys.argv[2] if len(sys.argv) > 2 else "catalog.json"
    extract_all_catalog_precios(p_dir, out_f)
