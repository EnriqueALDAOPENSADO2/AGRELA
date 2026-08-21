#!/usr/bin/env python3
import sys
import os
import json
import datetime
import openpyxl
from openpyxl.drawing.image import Image as OpenPyxlImage

def generate_invoice_excel(json_path, output_excel_path, template_path="sample.xlsx", logo_path="logo.jpg"):
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
        
    if not os.path.exists(template_path):
        template_path = os.path.join(os.path.dirname(__file__), "sample.xlsx")
        
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    # 1. Datos del Cliente
    cliente = data.get("cliente", {})
    ws["F5"].value = cliente.get("nombre", "")
    ws["F6"].value = cliente.get("direccion", "")
    ws["F7"].value = cliente.get("poblacion", "")
    ws["F8"].value = cliente.get("provincia", "")
    ws["F9"].value = cliente.get("cif_nif", "")
    
    # 2. Datos de Factura
    num_fac = str(data.get("numero_factura", "1"))
    ws["F11"].value = f"Nº {num_fac}" if not num_fac.startswith("Nº") else num_fac
    
    fecha_str = data.get("fecha", "")
    try:
        if fecha_str:
            d = datetime.date.fromisoformat(fecha_str)
            ws["H11"].value = d.strftime("%d/%m/%Y")
        else:
            ws["H11"].value = datetime.date.today().strftime("%d/%m/%Y")
    except Exception:
        ws["H11"].value = fecha_str
        
    # 3. Forma de Pago
    forma_pago = data.get("forma_pago", "TPV")
    ws["C40"].value = forma_pago
    
    # 4. Líneas de Artículos
    items = data.get("items", [])
    start_row = 13
    max_rows = 26 # filas 13 a 38
    
    for i in range(max_rows):
        r = start_row + i
        if i < len(items):
            it = items[i]
            desc = it.get("desc", "")
            uds = float(it.get("unidades", 1.0))
            precio = float(it.get("precio_unitario", 0.0))
            
            # Ancho Persiana Final -> se presenta en la factura
            ancho_final = float(it.get("ancho_persiana_final", 0.0))
            # Ancho Rollo Usado -> se usa para el cálculo
            ancho_rollo = float(it.get("ancho_rollo_usado", 0.0))
            alto = float(it.get("alto", 0.0))
            
            # Superficie M² calculada usando el ancho del rollo (o ancho final si rollo es 0)
            w_calc = ancho_rollo if ancho_rollo > 0 else ancho_final
            m2 = float(it.get("m2_calculado", 0.0))
            if m2 == 0.0 and w_calc > 0 and alto > 0:
                m2 = (w_calc / 1000.0) * (alto / 1000.0)
                if it.get("aplicar_minimo_compacto") and m2 < 1.50:
                    m2 = 1.50
            elif m2 == 0.0 and w_calc > 0 and alto == 0:
                m2 = (w_calc / 1000.0) # linear meters
                
            total_linea = float(it.get("total_calculado", 0.0))
            if total_linea == 0.0:
                total_linea = uds * precio * (m2 if m2 > 0 else 1.0)
                
            ws.cell(row=r, column=2).value = desc if desc else None
            ws.cell(row=r, column=5).value = uds
            ws.cell(row=r, column=6).value = precio
            ws.cell(row=r, column=7).value = ancho_final if ancho_final > 0 else None
            ws.cell(row=r, column=8).value = alto if alto > 0 else None
            ws.cell(row=r, column=9).value = round(m2, 3) if m2 > 0 else None
            ws.cell(row=r, column=10).value = round(total_linea, 2)
        else:
            # Limpiar fila vacía completamente
            ws.cell(row=r, column=2).value = None
            ws.cell(row=r, column=5).value = None
            ws.cell(row=r, column=6).value = None
            ws.cell(row=r, column=7).value = None
            ws.cell(row=r, column=8).value = None
            ws.cell(row=r, column=9).value = None
            ws.cell(row=r, column=10).value = None
            
    # 5. Totales
    total_bruto = sum(float(it.get("total_calculado", 0.0)) for it in items)
    if total_bruto == 0.0:
        total_bruto = float(data.get("total_bruto", 0.0))
        
    tipo_iva = float(data.get("tipo_iva", 0.21))
    cuota_iva = total_bruto * tipo_iva
    total_factura = total_bruto + cuota_iva
    
    ws["J39"].value = f"=SUM(J13:J38)"
    ws["G42"].value = tipo_iva
    ws["J42"].value = f"=((J39*{int(tipo_iva*100)})/100)"
    ws["J43"].value = f"=(J39+J42)"
    
    # 6. Pie de página Legal
    font_legal = openpyxl.styles.Font(name="Arial", size=7.5, italic=True, color="595959")
    ws["B45"].value = "PROTECCIÓN DE DATOS: Responsable: JUAN MANUEL ALDAO LOPEZ, 52434449S. Finalidad: Gestión de comunicaciones profesionales. Legitimación: Contrato e interés legítimo. Derechos y Más info: Puede ejercer sus derechos en mail persianasagrela@gmail.com."
    ws["B45"].font = font_legal
    ws["B46"].value = "CONFIDENCIALIDAD: Este mensaje es privado y dirigido solo al destinatario. La copia o difusión no autorizada está prohibida. Si lo recibe por error, por favor notifíquelo y elimínelo."
    ws["B46"].font = font_legal

    # 7. Logo
    actual_logo = logo_path if os.path.exists(logo_path) else ("logo.jpg" if os.path.exists("logo.jpg") else "logo.jpeg")
    if os.path.exists(actual_logo):
        try:
            # Solo añadir si no hay logo ya
            if len(ws._images) == 0:
                img = OpenPyxlImage(actual_logo)
                img.width = 160
                img.height = 110
                img.anchor = "B4"
                ws.add_image(img)
        except Exception as e:
            print(f"Nota: No se pudo adjuntar logo a Excel: {e}")
            
    wb.save(output_excel_path)
    print(f"Factura Excel guardada en: {output_excel_path}")
    return output_excel_path

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Uso: python export_invoice_excel.py <invoice_json> <output_xlsx> [template_xlsx] [logo_path]")
        sys.exit(1)
        
    json_p = sys.argv[1]
    out_xlsx = sys.argv[2]
    tmpl = sys.argv[3] if len(sys.argv) > 3 else "sample.xlsx"
    logo = sys.argv[4] if len(sys.argv) > 4 else "logo.jpg"
    
    generate_invoice_excel(json_p, out_xlsx, tmpl, logo)
