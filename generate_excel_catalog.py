import os
import re
import fitz  # PyMuPDF
import numpy as np
from PIL import Image as PILImage
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils.units import pixels_to_EMU

# Paths and Output configuration
OUTPUT_DIR = "extracted_images"
EXCEL_FILE = "AGRELA_Catalogo_Productos_2015.xlsx"
os.makedirs(OUTPUT_DIR, exist_ok=True)

def parse_price_and_unit(val_str, next_str=None):
    """Extracts numeric float price and unit string (e.g. 'ud.', 'ml.', 'm²')."""
    if not val_str:
        return None, ''
    val_str = str(val_str).strip()
    m = re.search(r'([0-9]+(?:[\.,][0-9]+)?)\s*([a-zA-Z²º\.]+)?', val_str)
    if not m:
        return None, ''
    num_str = m.group(1).replace('.', '').replace(',', '.') if ',' in m.group(1) else m.group(1)
    unit = m.group(2) or ''
    if not unit and next_str:
        next_clean = str(next_str).strip()
        if re.match(r'^[a-zA-Z²º\.]+$', next_clean):
            unit = next_clean
    try:
        price = float(num_str)
        return price, unit
    except ValueError:
        return None, unit

def crop_cell_image(page, cell_bbox, out_path):
    """
    Crops a table cell region at 300 DPI with a 1.5pt inset to exclude table border lines,
    auto-trims whitespace with a clean padding, and saves as a high-resolution PNG.
    """
    x0, y0, x1, y1 = cell_bbox
    rect = fitz.Rect(x0 + 1.5, y0 + 1.5, x1 - 1.5, y1 - 1.5)
    pix = page.get_pixmap(dpi=300, clip=rect)
    img = PILImage.frombytes('RGB', [pix.width, pix.height], pix.samples)
    
    arr = np.array(img)
    non_white = (arr[:, :, 0] < 240) | (arr[:, :, 1] < 240) | (arr[:, :, 2] < 240)
    rows = np.any(non_white, axis=1)
    cols = np.any(non_white, axis=0)
    
    if np.any(rows) and np.any(cols):
        ymin, ymax = np.where(rows)[0][[0, -1]]
        xmin, xmax = np.where(cols)[0][[0, -1]]
        pad = 8
        ymin = max(0, ymin - pad)
        ymax = min(arr.shape[0], ymax + pad)
        xmin = max(0, xmin - pad)
        xmax = min(arr.shape[1], xmax + pad)
        cropped_arr = arr[ymin:ymax, xmin:xmax]
        img = PILImage.fromarray(cropped_arr)
    else:
        return None
    
    img.save(out_path, format='PNG')
    return out_path

# ==========================================
# PARSER 1: COMPACTOS 2015
# ==========================================
def parse_compactos():
    doc = fitz.open('COMPACTOS 2015.pdf')
    page = doc[0]
    words = page.get_text('words')
    
    lines = {}
    for w in words:
        y = round(w[1], 1)
        m_y = next((ly for ly in lines if abs(ly - y) < 4.0), y)
        lines.setdefault(m_y, []).append(w)
    
    items = []
    notes = [
        'A efectos de cobro las medidas de ancho y alto se redondearán de 5 en 5 cms. cualquier medida intermedia se incrementará a la inmediata superior. Ejemplo: 132 x 143, se factura 135 x 145).',
        'El mínimo a facturar por unidad es de 1,50 m² .',
        'El IVA no está incluido en el precio y se cobrará aparte.'
    ]
    current_cat = ''
    
    for ly in sorted(lines.keys()):
        line_words = sorted(lines[ly], key=lambda x: x[0])
        txt = ' '.join([w[4] for w in line_words])
        
        if 'COMPACTO DE PVC' in txt:
            current_cat = txt.split('T-1')[0].strip()
            continue
            
        code_w = line_words[0][4]
        if code_w.isdigit() and len(code_w) == 9:
            rest_words = line_words[1:]
            p1_idx = None
            for idx, w in enumerate(rest_words):
                if re.match(r'^\d+,\d+$', w[4]):
                    p1_idx = idx
                    break
            if p1_idx is not None:
                desc = ' '.join([w[4] for w in rest_words[:p1_idx]])
                p1_val, u1 = parse_price_and_unit(rest_words[p1_idx][4], rest_words[p1_idx+1][4] if p1_idx+1 < len(rest_words) else None)
                p2_val, u2 = parse_price_and_unit(rest_words[p1_idx+2][4] if p1_idx+2 < len(rest_words) else None, rest_words[p1_idx+3][4] if p1_idx+3 < len(rest_words) else None)
                
                items.append({
                    'category': current_cat,
                    'code': code_w,
                    'desc': desc,
                    'p1': p1_val,
                    'u1': u1 or 'm²',
                    'p2': p2_val,
                    'u2': u2 or 'm²',
                    'img_path': None
                })
    return items, notes

# ==========================================
# PARSER 2: ACCESORIOS ENRROLLABLE Y COMPACTO
# ==========================================
def parse_accesorios():
    doc = fitz.open('Escandallo accesorios enrrollable y compacto.pdf')
    items = []
    
    for pno in range(4):
        page = doc[pno]
        tabs = page.find_tables()
        if not tabs.tables:
            continue
        tab = tabs.tables[0]
        extracted = tab.extract()
        
        current_img_path = None
        for r_idx, r in enumerate(tab.rows):
            if r_idx == 0:
                continue
            cell_0 = r.cells[0]
            row_vals = extracted[r_idx]
            
            code_w = (row_vals[1] or '').strip()
            if not code_w:
                continue
                
            if cell_0 is not None:
                img_name = f"accesorio_p{pno+1}_{code_w}.png"
                out_path = os.path.join(OUTPUT_DIR, img_name)
                current_img_path = crop_cell_image(page, cell_0, out_path)
            
            desc = (row_vals[2] or '').replace('\n', ' ').strip()
            p1_raw = row_vals[3] if len(row_vals) > 3 else None
            p2_raw = row_vals[4] if len(row_vals) > 4 else None
            
            p1_val, u1 = parse_price_and_unit(p1_raw)
            p2_val, u2 = parse_price_and_unit(p2_raw)
            
            items.append({
                'code': code_w,
                'desc': desc,
                'p1': p1_val,
                'u1': u1,
                'p2': p2_val,
                'u2': u2,
                'img_path': current_img_path
            })
    return items

# ==========================================
# PARSER 3: GUIAS Y PERFILES
# ==========================================
def parse_guias():
    doc = fitz.open('Guias y perfiles.pdf')
    items = []
    
    for pno in range(2):
        page = doc[pno]
        tabs = page.find_tables()
        for t_idx, tab in enumerate(tabs.tables):
            extracted = tab.extract()
            current_img_path = None
            for r_idx, r in enumerate(tab.rows):
                if r_idx == 0 and ('DISEÑO' in str(extracted[0]) or 'CODIGO' in str(extracted[0])):
                    continue
                row_vals = extracted[r_idx]
                code_w = (row_vals[1] or '').strip()
                desc = (row_vals[2] or '').replace('\n', ' ').strip()
                
                if not code_w and not desc:
                    continue
                if not code_w and not desc.startswith('Tapón'):
                    continue
                
                cell_0 = r.cells[0]
                if cell_0 is not None:
                    tag = code_w if code_w else desc.replace(' ', '_')
                    img_name = f"guia_p{pno+1}_t{t_idx}_{tag}.png"
                    out_path = os.path.join(OUTPUT_DIR, img_name)
                    current_img_path = crop_cell_image(page, cell_0, out_path)
                
                p1_raw = row_vals[3] if len(row_vals) > 3 else None
                p2_raw = row_vals[4] if len(row_vals) > 4 else None
                next_raw = row_vals[5] if len(row_vals) > 5 else None
                
                p1_val, u1 = parse_price_and_unit(p1_raw)
                p2_val, u2 = parse_price_and_unit(p2_raw, next_raw)
                
                items.append({
                    'code': code_w,
                    'desc': desc,
                    'p1': p1_val,
                    'u1': u1,
                    'p2': p2_val,
                    'u2': u2,
                    'img_path': current_img_path
                })
    return items

# ==========================================
# PARSER 4: TARIFAS MOTOR 2015
# ==========================================
def parse_motores():
    doc = fitz.open('TARIFAS MOTOR 2015.pdf')
    items = []
    notes = [
        '* En el precio del motor no está incluido ni soporte, coronas o adaptadores'
    ]
    
    for pno in range(2):
        page = doc[pno]
        tabs = page.find_tables()
        for t_idx, tab in enumerate(tabs.tables):
            extracted = tab.extract()
            current_img_path = None
            for r_idx, r in enumerate(tab.rows):
                if r_idx == 0 and ('DISEÑO' in str(extracted[0]) or 'CODIGO' in str(extracted[0])):
                    continue
                row_vals = extracted[r_idx]
                code_w = (row_vals[1] or '').strip()
                desc = (row_vals[2] or '').replace('\n', ' ').strip()
                if not code_w:
                    continue
                    
                cell_0 = r.cells[0]
                if cell_0 is not None:
                    img_name = f"motor_p{pno+1}_t{t_idx}_{code_w}.png"
                    out_path = os.path.join(OUTPUT_DIR, img_name)
                    current_img_path = crop_cell_image(page, cell_0, out_path)
                
                p1_raw = row_vals[3] if len(row_vals) > 3 else None
                p2_raw = row_vals[4] if len(row_vals) > 4 else None
                
                p1_val, u1 = parse_price_and_unit(p1_raw)
                p2_val, u2 = parse_price_and_unit(p2_raw)
                
                items.append({
                    'code': code_w,
                    'desc': desc,
                    'p1': p1_val,
                    'u1': u1,
                    'p2': p2_val,
                    'u2': u2,
                    'img_path': current_img_path
                })
    return items, notes

# ==========================================
# EXCEL BUILDER WITH STYLES & NUMERIC FORMATS
# ==========================================
def build_excel_catalog():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    # Styling Tokens
    HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ZEBRA_FILL = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    TITLE_FONT = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    SUBTITLE_FONT = Font(name="Calibri", size=11, italic=True, color="595959")
    CAT_FONT = Font(name="Calibri", size=12, bold=True, color="1F4E78")
    DATA_FONT = Font(name="Calibri", size=10)
    NOTE_FONT = Font(name="Calibri", size=10, italic=True, color="595959")
    
    ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    
    THIN_BORDER_SIDE = Side(style="thin", color="D9D9D9")
    BORDER_ALL = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=THIN_BORDER_SIDE)
    BORDER_HEADER = Border(left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE, top=THIN_BORDER_SIDE, bottom=Side(style="medium", color="1F4E78"))
    
    HEADERS = ["Imagen", "Código", "Producto / Descripción", "Precio T-1", "Unidad", "Precio T-1 / IVA", "Unidad IVA"]
    
    COL_A_WIDTH = 18    # approx 130px wide
    ROW_H_IMG = 55      # 55 pt = 73.33px tall
    ROW_H_NO_IMG = 24   # 24 pt
    
    CELL_W_PX = 130
    CELL_H_PX = 73
    MAX_IMG_W = 95
    MAX_IMG_H = 52
    
    def create_sheet_structure(sheet_title, main_title, items, notes=None):
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        
        # Title Block
        ws.merge_cells("A1:G1")
        ws["A1"] = main_title
        ws["A1"].font = TITLE_FONT
        ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28
        
        ws.merge_cells("A2:G2")
        ws["A2"] = "AGRELA - Tarifa de Precios (Enero 2015) | IVA 21%"
        ws["A2"].font = SUBTITLE_FONT
        ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[2].height = 20
        
        # Header Row
        header_row = 4
        ws.row_dimensions[header_row].height = 26
        for col_num, header in enumerate(HEADERS, 1):
            cell = ws.cell(row=header_row, column=col_num, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = ALIGN_CENTER
            cell.border = BORDER_HEADER
            
        current_row = 5
        last_cat = None
        
        for idx, item in enumerate(items):
            cat = item.get('category')
            if cat and cat != last_cat:
                last_cat = cat
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                cat_cell = ws.cell(row=current_row, column=1, value=cat)
                cat_cell.fill = SUBHEADER_FILL
                cat_cell.font = CAT_FONT
                cat_cell.alignment = ALIGN_LEFT
                ws.row_dimensions[current_row].height = 25
                for c in range(1, 8):
                    ws.cell(row=current_row, column=c).border = BORDER_ALL
                current_row += 1
                
            has_img = bool(item.get('img_path') and os.path.exists(item['img_path']))
            ws.row_dimensions[current_row].height = ROW_H_IMG if has_img else ROW_H_NO_IMG
            fill = ZEBRA_FILL if idx % 2 == 1 else WHITE_FILL
            
            # Col A: Imagen
            cell_a = ws.cell(row=current_row, column=1)
            cell_a.fill = fill
            cell_a.alignment = ALIGN_CENTER
            cell_a.border = BORDER_ALL
            
            if has_img:
                try:
                    img = OpenPyxlImage(item['img_path'])
                    w, h = img.width, img.height
                    ratio = min(float(MAX_IMG_W) / w, float(MAX_IMG_H) / h, 1.0)
                    target_w = max(1, int(w * ratio))
                    target_h = max(1, int(h * ratio))
                    img.width = target_w
                    img.height = target_h
                    
                    offset_x_px = max(0, (CELL_W_PX - target_w) // 2)
                    offset_y_px = max(0, (CELL_H_PX - target_h) // 2)
                    
                    marker = AnchorMarker(
                        col=0,
                        colOff=pixels_to_EMU(offset_x_px),
                        row=current_row - 1,
                        rowOff=pixels_to_EMU(offset_y_px)
                    )
                    size = XDRPositiveSize2D(pixels_to_EMU(target_w), pixels_to_EMU(target_h))
                    img.anchor = OneCellAnchor(_from=marker, ext=size)
                    ws.add_image(img)
                except Exception as e:
                    print(f"Error adding image {item['img_path']}: {e}")
            
            # Col B: Código
            cell_b = ws.cell(row=current_row, column=2, value=item['code'])
            cell_b.fill = fill
            cell_b.font = DATA_FONT
            cell_b.alignment = ALIGN_CENTER
            cell_b.border = BORDER_ALL
            cell_b.number_format = '@'
            
            # Col C: Producto / Descripción
            cell_c = ws.cell(row=current_row, column=3, value=item['desc'])
            cell_c.fill = fill
            cell_c.font = DATA_FONT
            cell_c.alignment = ALIGN_LEFT
            cell_c.border = BORDER_ALL
            
            # Col D: Precio T-1
            cell_d = ws.cell(row=current_row, column=4)
            cell_d.fill = fill
            cell_d.font = DATA_FONT
            cell_d.alignment = ALIGN_RIGHT
            cell_d.border = BORDER_ALL
            if item['p1'] is not None:
                cell_d.value = float(item['p1'])
                cell_d.number_format = '#,##0.00 "€"'
                
            # Col E: Unidad
            cell_e = ws.cell(row=current_row, column=5, value=item['u1'])
            cell_e.fill = fill
            cell_e.font = DATA_FONT
            cell_e.alignment = ALIGN_CENTER
            cell_e.border = BORDER_ALL
            
            # Col F: Precio T-1 / IVA
            cell_f = ws.cell(row=current_row, column=6)
            cell_f.fill = fill
            cell_f.font = DATA_FONT
            cell_f.alignment = ALIGN_RIGHT
            cell_f.border = BORDER_ALL
            if item['p2'] is not None:
                cell_f.value = float(item['p2'])
                cell_f.number_format = '#,##0.00 "€"'
                
            # Col G: Unidad IVA
            cell_g = ws.cell(row=current_row, column=7, value=item['u2'])
            cell_g.fill = fill
            cell_g.font = DATA_FONT
            cell_g.alignment = ALIGN_CENTER
            cell_g.border = BORDER_ALL
            
            current_row += 1
            
        # Add Notes if present
        if notes:
            current_row += 1
            for note in notes:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=7)
                note_cell = ws.cell(row=current_row, column=1, value=f"Nota: {note}" if not note.startswith('*') else note)
                note_cell.font = NOTE_FONT
                note_cell.alignment = ALIGN_LEFT
                ws.row_dimensions[current_row].height = 20
                current_row += 1
                
        # Set Column Widths
        column_widths = {
            'A': COL_A_WIDTH,
            'B': 14,
            'C': 55,
            'D': 15,
            'E': 10,
            'F': 16,
            'G': 12,
        }
        for col_letter, width in column_widths.items():
            ws.column_dimensions[col_letter].width = width
            
        # Enable AutoFilter
        ws.auto_filter.ref = f"A4:G{current_row-1}"

    # Build Sheet 1: Compactos
    print("Processing Compactos...")
    compactos_items, compactos_notes = parse_compactos()
    create_sheet_structure("Compactos", "CATÁLOGO DE COMPACTOS DE PVC 155 Y 185", compactos_items, compactos_notes)
    
    # Build Sheet 2: Accesorios
    print("Processing Accesorios...")
    accesorios_items = parse_accesorios()
    create_sheet_structure("Accesorios", "ESCANDALLO ACCESORIOS ENROLLABLE Y COMPACTO", accesorios_items)
    
    # Build Sheet 3: Guías y Perfiles
    print("Processing Guías y Perfiles...")
    guias_items = parse_guias()
    create_sheet_structure("Guías y Perfiles", "TARIFA DE PRECIOS - GUÍAS Y PERFILES", guias_items)
    
    # Build Sheet 4: Motores y Automatismos
    print("Processing Motores y Automatismos...")
    motores_items, motores_notes = parse_motores()
    create_sheet_structure("Motores y Automatismos", "MOTORES, AUTOMATISMOS Y ACCESORIOS", motores_items, motores_notes)

    wb.save(EXCEL_FILE)
    print(f"Excel catalog successfully updated and saved at {EXCEL_FILE}")

if __name__ == "__main__":
    build_excel_catalog()
